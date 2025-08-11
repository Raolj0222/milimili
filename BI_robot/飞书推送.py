import pandas as pd
import sys
# path为tools的上层文件夹
sys.path.append('D:\milimili')
from tools.mysqlhandle import Mysql_handle
import datetime
import openpyxl
from configs import mili_mysql
mysql = Mysql_handle(mili_mysql,'data_analysis')
from get_img import img_get
from wangzuan import app_info
from feishu import FeiShu,web_hook_test,web_hook,web_hook_new
feishu = FeiShu()
# 设置开始时间和结束时间
end = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
start = (datetime.date.today() - datetime.timedelta(days=7)).strftime('%Y-%m-%d')


wangzhuan_df,cp_df=app_info(start,end)
# 根据日期去重
wangzhuan_df = wangzhuan_df.drop_duplicates(subset=['date','appname','country'],keep = 'first')
cp_df = cp_df.drop_duplicates(subset=['date','appname','country'],keep='first')
# print(cp_df)
# 筛选出投放数据
# wangzhuan_df=wangzhuan_df[wangzhuan_df['money']==wangzhuan_df['money']]
# cp_df=cp_df[cp_df['money']==cp_df['money']]
wangzhuan_df.fillna(value=0,inplace=True)
cp_df.fillna(value=0,inplace=True)
# 处理需要计算的列
# 网赚
wangzhuan_list=['date','appname','country','总支出','总收入','总ROI','毛利','总体ARPU',
                'cost_money','新增成本','pushinstall','money','推广CPI','install','daus','新增活跃比','自然新增',
                'averageTime_PerDevice','广告收入','广告ROI','广告ARPU','激励人均展示','插页人均展示','banner人均展示',
                '激励收入','插页收入','banner收入','激励ecpm','插页ecpm','bannerecpm','积分墙收入','积分墙ROI',
                '积分墙ARPU','adjoe收入','adjoeROI','adjoe_ecpm','okspin收入','okspinROI','okspin_ecpm'
                ]
# cp产品
cp_list=['date','appname','country','总支出','总收入','总ROI','毛利','总体ARPU',
                '新增成本','pushinstall','money','推广CPI','install','daus','活跃付费用户','自然新增',
                'averageTime_PerDevice','广告收入','广告ROI','广告ARPU','激励人均展示','插页人均展示','banner人均展示',
                '激励收入','插页收入','banner收入','激励ecpm','插页ecpm','bannerecpm',
                'ng','ng_users','付费ROI','付费率','付费arpu','新用户收入','老用户收入',
                '小R用户收入','中R用户收入','大R用户收入','积分墙收入','积分墙ROI']
cp_list_new=['date','appname','country','总支出','总收入','总ROI','总ROI(扣除分成)','毛利','毛利(扣除分成)','总体ARPU',
                '新增成本','pushinstall','money','推广CPI','install','daus','活跃付费用户','自然新增',
                'averageTime_PerDevice','广告收入','广告ROI','广告ARPU','激励人均展示','插页人均展示','banner人均展示',
                '激励收入','插页收入','banner收入','激励ecpm','插页ecpm','bannerecpm',
                'ng','ng_users','付费ROI','付费率','付费arpu','新用户收入','老用户收入',
                '小R用户收入','中R用户收入','大R用户收入','积分墙收入','积分墙ROI']
# 写入excel表格
# keno 和 stars 停止投放 daub farm 目前是余量收益，无投放
wzngahzn_app_list={
    'Daub Farm':  ['US']
}
cp_app_list={
    'Doomsday Vanguard': ['all'],
    'Doomsday Vanguard ios': ['all'],
    # 'Tales of Brave':['JP','all']
    # 'Bingo Party': ['US'],
    # 'Keno Farm B':['BR']
}
cp_app_new_list={
'Tales of Brave':['JP','all'],
'Tales of Brave ios':['JP','all'],

}

wangzhuan_file=r'D:\milimili\BI_robot\网赚报表.xlsx'
cp_file= r'D:\milimili\BI_robot\cp报表.xlsx'
cp_new_file=r'D:\milimili\BI_robot\cp报表扣除分成.xlsx'
def wangzhuan_post():
    # wangzhuan_df,cp_df=app_info(start,end)
    # # 筛选出投放数据
    # # wangzhuan_df = wangzhuan_df[wangzhuan_df['money'] == wangzhuan_df['money']]
    # # cp_df = cp_df[cp_df['money'] == cp_df['money']]
    # wangzhuan_df.fillna(value=0, inplace=True)
    # cp_df.fillna(value=0, inplace=True)
    # 处理需要计算的列
    # 网赚
    wangzhuan_df['总支出'] = wangzhuan_df['cost_money'] + wangzhuan_df['money']
    wangzhuan_df['广告收入'] = wangzhuan_df['激励收入'] + wangzhuan_df['插页收入']
    wangzhuan_df['广告ROI'] = wangzhuan_df['广告收入'] / wangzhuan_df['总支出']
    wangzhuan_df['广告ARPU'] = wangzhuan_df['广告收入'] / wangzhuan_df['daus']
    wangzhuan_df['激励人均展示'] = wangzhuan_df['激励展示'] / wangzhuan_df['daus']
    wangzhuan_df['插页人均展示'] = wangzhuan_df['插页展示'] / wangzhuan_df['daus']
    wangzhuan_df['banner人均展示'] = wangzhuan_df['banner展示'] / wangzhuan_df['daus']
    wangzhuan_df['激励ecpm'] = wangzhuan_df['激励收入'] / wangzhuan_df['激励展示'] * 1000
    wangzhuan_df['插页ecpm'] = wangzhuan_df['插页收入'] / wangzhuan_df['插页展示'] * 1000
    wangzhuan_df['bannerecpm'] = wangzhuan_df['banner收入'] / wangzhuan_df['banner展示'] * 1000
    wangzhuan_df['积分墙收入'] = wangzhuan_df['adjoe收入'] + wangzhuan_df['okspin收入']
    wangzhuan_df['积分墙ROI'] = wangzhuan_df['积分墙收入'] / wangzhuan_df['总支出']
    wangzhuan_df['积分墙ARPU'] = wangzhuan_df['积分墙收入'] / wangzhuan_df['daus']
    wangzhuan_df['adjoeROI'] = wangzhuan_df['adjoe收入'] / wangzhuan_df['总支出']
    wangzhuan_df['okspinROI'] = wangzhuan_df['okspin收入'] / wangzhuan_df['总支出']
    wangzhuan_df['总收入'] = wangzhuan_df['广告收入'] + wangzhuan_df['积分墙收入']
    wangzhuan_df['总ROI'] = wangzhuan_df['总收入'] / wangzhuan_df['总支出']
    wangzhuan_df['毛利'] = wangzhuan_df['总收入'] - wangzhuan_df['总支出']
    wangzhuan_df['总体ARPU'] = wangzhuan_df['总收入'] / wangzhuan_df['daus']
    wangzhuan_df['新增成本'] = wangzhuan_df['money'] / wangzhuan_df['install']
    wangzhuan_df['推广CPI'] = wangzhuan_df['money'] / wangzhuan_df['pushinstall']
    wangzhuan_df['新增活跃比'] = wangzhuan_df['daus'] / wangzhuan_df['install']
    wangzhuan_df['自然新增'] = wangzhuan_df['install'] - wangzhuan_df['pushinstall']

    wangzhuan_df.to_excel('wangzhuan.xlsx')

    for app,country in wzngahzn_app_list.items():
        # 筛选app和地区
        for country_code in country:
            res = wangzhuan_df[(wangzhuan_df['appname']==f'{app}')&(wangzhuan_df['country']==f'{country_code}')]
            res.sort_values(by='date',ascending=True,inplace=True)
            # 开始写入,按照表格顺序依次写入
            column_num = 0
            for column in wangzhuan_list:
                column_num = 1 + column_num
                wb = openpyxl.load_workbook(wangzhuan_file)
                ws = wb['sheet1']
                values = res[column].tolist()
                for i in range(0,len(values)):
                    value=values[i]
                    # 从表格的第三行开始写入
                    ws.cell(row=i + 3, column=column_num).value = value
                wb.save(wangzhuan_file)
                wb.close()
            #每一个国家写入完成后开始截图
            # 设置文件名
            p_name='wzimg'
            print(p_name)
            img_get(wangzhuan_file,'sheet1',p_name)

            # 截图完成开始推送
            # 设置富文本标题：
            roi=round(res[res['date']==end]['总ROI'].tolist()[0]*100,1)
            # print(roi)
            maoli=round(res[res['date']==end]['毛利'].tolist()[0],2)
            title=f'{app}-{country_code}：昨日总ROI:{roi}%,毛利:{maoli}'
            print(title)
            # 发送至飞书群
            feishu.run(title,'D:\milimili\BI_robot\截图-网赚报表\wzimg.png',web_hook_test)
def cp_post(file,cplist):
    # wangzhuan_df, cp_df = app_info(start, end)
    # # cp_df = cp_df[cp_df['money'] == cp_df['money']]
    # cp_df.fillna(value=0, inplace=True)
    cp_df['总支出'] = cp_df['money']
    cp_df['广告收入'] = cp_df['激励收入'] + cp_df['插页收入']
    cp_df['广告ROI'] = cp_df['广告收入'] / cp_df['总支出']
    cp_df['广告ARPU'] = cp_df['广告收入'] / cp_df['daus']
    cp_df['激励人均展示'] = cp_df['激励展示'] / cp_df['daus']
    cp_df['插页人均展示'] = cp_df['插页展示'] / cp_df['daus']
    cp_df['banner人均展示'] = cp_df['banner展示'] / cp_df['daus']
    cp_df['激励ecpm'] = cp_df['激励收入'] / cp_df['激励展示'] * 1000
    cp_df['插页ecpm'] = cp_df['插页收入'] / cp_df['插页展示'] * 1000
    cp_df['bannerecpm'] = cp_df['banner收入'] / cp_df['banner展示'] * 1000
    cp_df['总收入'] = cp_df['广告收入'] + cp_df['ng'] + cp_df['积分墙收入']
    cp_df['总ROI'] = cp_df['总收入'] / cp_df['总支出']
    cp_df['毛利'] = cp_df['总收入'] - cp_df['总支出']
    cp_df['总体ARPU'] = cp_df['总收入'] / cp_df['daus']
    cp_df['新增成本'] = cp_df['money'] / cp_df['install']
    cp_df['推广CPI'] = cp_df['money'] / cp_df['pushinstall']
    cp_df['新增活跃比'] = cp_df['daus'] / cp_df['install']
    cp_df['活跃付费用户'] = cp_df['users']
    cp_df['自然新增'] = cp_df['install'] - cp_df['pushinstall']
    cp_df['付费ROI'] = cp_df['ng'] / cp_df['总支出']
    cp_df['付费率'] = cp_df['ng_users'] / cp_df['daus']
    cp_df['付费arpu']=cp_df['ng'] / cp_df['daus']
    cp_df['积分墙ROI'] = cp_df['积分墙收入'] / cp_df['总支出']
    # print(cp_df.query('appname=="Keno Farm B"&country=="BR"')[['积分墙收入','积分墙ROI']])

    for app, country in cplist.items():
        # 筛选app和地区
        for country_code in country:
            res = cp_df[(cp_df['appname'] == f'{app}') & (cp_df['country'] == f'{country_code}')]
            res.sort_values(by='date',ascending=True,inplace=True)
            print(res)
            # 开始写入,按照表格顺序依次写入
            column_num = 0
            for column in cp_list:
                column_num = 1 + column_num
                wb = openpyxl.load_workbook(cp_file)
                ws = wb['sheet1']
                values = res[column].tolist()
                for i in range(0, len(values)):
                    value = values[i]
                    # 从表格的第三行开始写入
                    ws.cell(row=i + 3, column=column_num).value = value
                wb.save(cp_file)
                wb.close()
            # 每一个国家写入完成后开始截图
            # 设置文件名
            p_name = 'cpimg'
            print(p_name)
            img_get(cp_file, 'sheet1', p_name)
            # 截图完成开始推送
            # 设置富文本标题：
            roi = round(res[res['date'] == end]['总ROI'].tolist()[0]* 100, 1)
            maoli = round(res[res['date'] == end]['毛利'].tolist()[0], 2)
            title = f'{app} - {country_code}：昨日总ROI: {roi}%,毛利: {maoli}'
            print(title)
            # 发送至飞书群
            feishu.run(title, 'D:\milimili\BI_robot\截图-cp报表\cpimg.png', file)
def cp_post_new():
    # wangzhuan_df, cp_df = app_info(start, end)
    # # cp_df = cp_df[cp_df['money'] == cp_df['money']]
    # cp_df.fillna(value=0, inplace=True)
    cp_df['总支出'] = cp_df['money']
    cp_df['广告收入'] = cp_df['激励收入'] + cp_df['插页收入']
    cp_df['广告ROI'] = cp_df['广告收入'] / cp_df['总支出']
    cp_df['广告ARPU'] = cp_df['广告收入'] / cp_df['daus']
    cp_df['激励人均展示'] = cp_df['激励展示'] / cp_df['daus']
    cp_df['插页人均展示'] = cp_df['插页展示'] / cp_df['daus']
    cp_df['banner人均展示'] = cp_df['banner展示'] / cp_df['daus']
    cp_df['激励ecpm'] = cp_df['激励收入'] / cp_df['激励展示'] * 1000
    cp_df['插页ecpm'] = cp_df['插页收入'] / cp_df['插页展示'] * 1000
    cp_df['bannerecpm'] = cp_df['banner收入'] / cp_df['banner展示'] * 1000
    cp_df['总收入'] = cp_df['广告收入'] + cp_df['ng'] + cp_df['积分墙收入']
    cp_df['总ROI'] = cp_df['总收入'] / cp_df['总支出']
    cp_df['总ROI(扣除分成)'] = (cp_df['总收入']*0.88) / cp_df['总支出']
    cp_df['毛利'] = cp_df['总收入'] - cp_df['总支出']
    cp_df['毛利(扣除分成)'] = (cp_df['总收入']*0.88) - cp_df['总支出']
    cp_df['总体ARPU'] = cp_df['总收入'] / cp_df['daus']
    cp_df['新增成本'] = cp_df['money'] / cp_df['install']
    cp_df['推广CPI'] = cp_df['money'] / cp_df['pushinstall']
    cp_df['新增活跃比'] = cp_df['daus'] / cp_df['install']
    cp_df['活跃付费用户'] = cp_df['users']
    cp_df['自然新增'] = cp_df['install'] - cp_df['pushinstall']
    cp_df['付费ROI'] = cp_df['ng'] / cp_df['总支出']
    cp_df['付费率'] = cp_df['ng_users'] / cp_df['daus']
    cp_df['付费arpu']=cp_df['ng'] / cp_df['daus']
    cp_df['积分墙ROI'] = cp_df['积分墙收入'] / cp_df['总支出']
    # print(cp_df.query('appname=="Keno Farm B"&country=="BR"')[['积分墙收入','积分墙ROI']])

    for app, country in cp_app_new_list.items():
        # 筛选app和地区
        for country_code in country:
            res = cp_df[(cp_df['appname'] == f'{app}') & (cp_df['country'] == f'{country_code}')]
            # print(res.columns)
            res.sort_values(by='date',ascending=True,inplace=True)
            # 开始写入,按照表格顺序依次写入
            column_num = 0
            for column in cp_list_new:
                column_num = 1 + column_num
                wb = openpyxl.load_workbook(cp_new_file)
                ws = wb['sheet1']
                values = res[column].tolist()
                for i in range(0, len(values)):
                    value = values[i]
                    # 从表格的第三行开始写入
                    ws.cell(row=i + 3, column=column_num).value = value
                wb.save(cp_new_file)
                wb.close()
            # 每一个国家写入完成后开始截图
            # 设置文件名
            p_name = 'cpimg_new'
            print(p_name)
            img_get(cp_new_file, 'sheet1', p_name)
            # 截图完成开始推送
            # 设置富文本标题：
            roi = round(res[res['date'] == end]['总ROI'].tolist()[0]* 100, 1)
            roi_new = round(res[res['date'] == end]['总ROI(扣除分成)'].tolist()[0]* 100, 1)
            maoli = round(res[res['date'] == end]['毛利'].tolist()[0], 2)
            maoli_new = round(res[res['date'] == end]['毛利(扣除分成)'].tolist()[0], 2)
            title = f'{app} - {country_code}：昨日总ROI: {roi}%,扣除分成后ROI：{roi_new}%,毛利: {maoli},扣除分成后毛利：{maoli_new}'
            print(title)
            # 发送至飞书群
            feishu.run(title, 'D:\milimili\BI_robot\截图-cp报表扣除分成\cpimg_new.png', web_hook_test)



if __name__ == '__main__':
    # wangzhuan_post()
    cp_post(web_hook_test,cp_app_list)# 滋滋日报
    # cp_post(web_hook_test,cp_app_new_list) # 勇者日报 cp专用
    cp_post_new() # 勇者日报特殊处理（团队自用）
    pass
# # 设置定时推送
# from apscheduler.schedulers.blocking import BlockingScheduler
# # 参数说明 https://blog.csdn.net/lipachong/article/details/99962134
# def my_schedule():
#             schedu = BlockingScheduler()
#             schedu.add_job(post_feishu(), 'cron',
# 			  day_of_week='0-6',# 表示周一到周天
# 			  hour='10', minute='50', second='00')
#             schedu.start()




